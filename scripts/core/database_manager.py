"""
数据库管理器
处理Excel数据库的读写操作
"""
import os,sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Any, Tuple
import shutil
from datetime import datetime
import warnings
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../'))

# 导入配置和模型
from scripts.core.config_loader import get_config_instance
from scripts.core.database_model import Paper, is_same_identity, is_duplicate_paper
from scripts.utils import normalize_dataframe_columns  # 新增导入


class DatabaseManager:
    """数据库管理器"""
    
    def __init__(self):
        self.config = get_config_instance()
        self.settings = get_config_instance().settings
        self.core_excel_path = self.settings['paths']['core_excel']
        self.backup_dir = self.settings['paths']['backup_dir']
        self.conflict_marker = self.settings['database']['conflict_marker']
        
        # 确保目录存在
        os.makedirs(os.path.dirname(self.core_excel_path), exist_ok=True)
        os.makedirs(self.backup_dir, exist_ok=True)
    
    def backup_database(self) -> str:
        """备份数据库，返回备份文件路径"""
        if not os.path.exists(self.core_excel_path):
            return ""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"paper_database_backup_{timestamp}.xlsx"
        backup_path = os.path.join(self.backup_dir, backup_filename)
        
        shutil.copy2(self.core_excel_path, backup_path)
        print(f"数据库已备份到: {backup_path}")
        return backup_path
    
    def load_database(self) -> pd.DataFrame:
        """加载数据库到DataFrame"""
        if not os.path.exists(self.core_excel_path):
            # 创建新的数据库文件
            return self._create_new_database()
        
        try:
            # 尝试读取Excel文件
            df = pd.read_excel(self.core_excel_path, engine='openpyxl')
            
            # 确保所有必需的列都存在
            return self._ensure_columns_exist(df)
        except Exception as e:
            print(f"加载数据库失败: {e}")
            # 创建新的数据库文件
            return self._create_new_database()
    
    def _create_new_database(self) -> pd.DataFrame:
        """创建新的数据库DataFrame"""
        # 获取所有激活的标签
        active_tags = self.config.get_active_tags()
        
        # 创建空的DataFrame，按照order排序
        active_tags.sort(key=lambda x: x['order'])
        columns = [tag['table_name'] for tag in active_tags]
        
        df = pd.DataFrame(columns=columns)
        print(f"创建新的数据库文件: {self.core_excel_path}")
        return df
    
    def _ensure_columns_exist(self, df: pd.DataFrame) -> pd.DataFrame:
        """确保DataFrame包含所有必需的列"""
        active_tags = self.config.get_active_tags()
        active_tags.sort(key=lambda x: x['order'])
        required_columns = [tag['table_name'] for tag in active_tags]
        
        # 添加缺失的列
        for column in required_columns:
            if column not in df.columns:
                df[column] = None
        
        # 移除多余的列（不在激活标签中的列）
        columns_to_keep = [col for col in df.columns if col in required_columns]
        
        # 重新排序列
        df = df[required_columns]
        
        return df
    
    def save_database(self, df: pd.DataFrame, password: str = "") -> bool:
        """保存DataFrame到Excel文件"""
        try:
            # 先根据 tag_config 规范 DataFrame 列与 category 值
            df = normalize_dataframe_columns(df, self.config)

            # 备份原文件
            self.backup_database()
            
            # 保存到Excel
            with pd.ExcelWriter(
                self.core_excel_path,
                engine='openpyxl'
            ) as writer:
                df.to_excel(writer, index=False, sheet_name='Papers')
                
                # 获取工作簿和工作表
                workbook = writer.book
                worksheet = writer.sheets['Papers']
                
                # 应用格式
                self._apply_excel_formatting(workbook, worksheet, df)
                
                # 如果有密码，尝试设置保护（但openpyxl的写保护有限）
                if password:
                    try:
                        worksheet.protection.set_password(password)
                        worksheet.protection.sheet = True
                    except:
                        print("注意：无法设置Excel保护，文件将以未加密形式保存")
            
            print(f"数据库已保存到: {self.core_excel_path}")
            return True
        except Exception as e:
            print(f"保存数据库失败: {e}")
            return False
    
    def _apply_excel_formatting(self, workbook, worksheet, df):
        """应用Excel格式"""
        # 设置列宽
        for column in df.columns:
            column_letter = get_column_letter(df.columns.get_loc(column) + 1)
            max_length = max(
                df[column].astype(str).map(len).max(),
                len(str(column))
            )
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 设置标题行样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # 标记冲突行
        conflict_row_name=self.config.get_tag_field("conflict_marker","table_name")
        if conflict_row_name in df.columns:
            conflict_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            for idx, row in df.iterrows():
                if row.get(conflict_row_name).strip() not in [False,0,None,"False","FALSE","false","","0"]:
                    for cell in worksheet[idx + 2]:  # +2因为标题行是1，索引从0开始
                        cell.fill = conflict_fill
                        #如果列名是title，就在该cell内容前加冲突标记
                        if cell.column_letter == get_column_letter(df.columns.get_loc('title') + 1):
                            cell.value = f"{row['conflict_marker']} {cell.value}"
        else:
            print(f"添加论文冲突格式时，发现数据库表格没有conflict_marker列")

                        
    
    def get_password(self) -> str:
        """获取Excel密码"""
        # 首先尝试从环境变量获取
        password = os.environ.get('EXCEL_KEY', '')
        if password:
            return password
        
        # 尝试从本地文件获取
        password_path = self.settings['excel'].get('password_path', '')
        if password_path and os.path.exists(password_path):
            try:
                with open(password_path, 'r', encoding='utf-8') as f:
                    password = f.read().strip()
                    return password
            except Exception as e:
                print(f"本地读取Excel密码失败{e}")
                
        
        return ""  # 返回空字符串表示不加密
    
    def get_papers_from_dataframe(self, df: pd.DataFrame) -> List[Paper]:
        """从DataFrame提取Paper对象列表"""
        papers = []
        
        for _, row in df.iterrows():
            paper_data = {}
            
            # 将DataFrame行转换为字典
            for tag in self.config.get_active_tags():
                column_name = tag['table_name']
                if column_name in row:
                    value = row[column_name]
                    
                    # 处理NaN值 -> 统一规范化
                    if pd.isna(value):
                        value = ""
                    
                    # 根据标签类型进行合理转换：
                    # - bool/int 保持类型（尽量恢复原始语义）
                    # - 其它全部转换为 str 并去除首尾空白
                    t = tag.get('type', 'string')
                    if t == 'bool':
                        # 空字符串视为 False
                        value =True if value not in (False,0,"", None,"false","FALSE","False","0") else False
                    elif t == 'int':
                        try:
                            value = int(value) if value not in ("", None) else 0
                        except Exception:
                            value = 0
                    else:
                        # 强制转为 string
                        value = str(value).strip() if value is not None else ""
                    
                    paper_data[tag['variable']] = value
            
            # 创建Paper对象
            paper = Paper.from_dict(paper_data)
            papers.append(paper)
        
        return papers
    
    def get_dataframe_from_papers(self, papers: List[Paper]) -> pd.DataFrame:
        """从Paper对象列表创建DataFrame"""
        # 获取激活的标签
        active_tags = self.config.get_active_tags()
        active_tags.sort(key=lambda x: x['order'])
        
        # 准备数据
        data = []
        for paper in papers:
            row = {}
            for tag in active_tags:
                column_name = tag['table_name']
                value = getattr(paper, tag['variable'], "")
                row[column_name] = value
            data.append(row)
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 确保列顺序正确
        column_order = [tag['table_name'] for tag in active_tags]
        df = df[column_order]
        
        return df
    
    #唯一对外接口
    def add_papers(self, new_papers: List[Paper], conflict_resolution: str = 'mark') -> Tuple[List[Paper], List[Paper]]:
        """
        添加新论文到数据库，同时验证冲突
        
        参数:
            new_papers: 新论文列表
            conflict_resolution: 冲突解决策略 ('mark', 'skip', 'replace')
        
        返回:
            Tuple[成功添加的论文列表, 冲突论文列表（被标记后需要加入数据库的）]
        """
        # 加载现有数据库
        df = self.load_database()
        existing_papers = self.get_papers_from_dataframe(df)
        
        added_papers = []
        conflict_papers = []

        for new_paper in new_papers:

            # 检查是否已存在（按 DOI 或 title 判断同一论文，忽略 conflict_marker）
            is_same_paper = False
            duplicate_papers = []
            conflict_with = None
            conflict_index = 0
            
            # 找出所有同identity论文
            for idx, existing_paper in enumerate(existing_papers):
                if is_same_identity(new_paper, existing_paper):
                    is_same_paper = True
                    duplicate_papers.append(existing_paper)
                    # 唯一一个最初版本
                    if existing_paper.conflict_marker == False:
                        conflict_with = existing_paper
                        conflict_index = idx

            if is_same_paper:
                if conflict_with == None:
                    print("提交时发现存在同指向论文，但无法找到最初论文，已停止写入新论文，请检查")
                    return [], new_papers
                
                # 如果是同一论文身份，先判断是否为"完全重复提交"
                if is_duplicate_paper(duplicate_papers, new_paper):
                    # 完全相同，跳过
                    print(f"论文: {new_paper.title}重复提交，跳过添加")
                    continue

                # 不是完全相同，按冲突策略处理
                if conflict_resolution == 'skip':
                    continue
                # 完全替换
                elif conflict_resolution == 'replace':
                    existing_papers = [p for p in existing_papers if not is_same_identity(p, new_paper)]
                    # 这里不直接插入，而是先添加到列表末尾，稍后排序
                    existing_papers.append(new_paper)
                    added_papers.append(new_paper)
                elif conflict_resolution == 'mark':
                    new_paper.conflict_marker = True
                    conflict_papers.append((new_paper, conflict_with))
                    # 不直接插入，稍后排序处理
                    existing_papers.append(new_paper)
                    added_papers.append(new_paper)
            else:
                # 新论文，添加到列表末尾，稍后排序
                existing_papers.append(new_paper)
                added_papers.append(new_paper)

        # 修正的排序逻辑：
        # 1. 首先找出所有无冲突标记的paper
        non_conflict_papers = [p for p in existing_papers if not p.conflict_marker]
        conflict_marked_papers = [p for p in existing_papers if p.conflict_marker]
        
        # 2. 对无冲突标记的paper按category排序，category内部越晚提交的越靠前
        # 需要为paper添加一个提交时间戳属性（假设为submission_time）
        # 这里假设paper有submission_time属性，如果没有需要添加
        
        # 按category分组排序
        category_groups = {}
        for paper in non_conflict_papers:
            if paper.category not in category_groups:
                category_groups[paper.category] = []
            category_groups[paper.category].append(paper)
        
        # 每个category内部按提交时间倒序（越晚提交越靠前）
        sorted_non_conflict = []
        for category in sorted(category_groups.keys()):
            papers_in_category = category_groups[category]
            # 按提交时间倒序排序（假设有submission_time属性）
            papers_in_category.sort(key=lambda x: x.submission_time, reverse=True)
            sorted_non_conflict.extend(papers_in_category)
        
        # 3. 处理冲突标记的paper：找到最早的无冲突标的paper，放在其正上方
        # 越晚提交的冲突标记paper越靠近这个最早的无冲突标的paper
        
        # 首先按提交时间对冲突标记paper排序（越晚提交越靠后，因为插入时从后往前）
        if conflict_marked_papers:
            # 假设有submission_time属性
            conflict_marked_papers.sort(key=lambda x: x.submission_time)
            
            # 找到最早的无冲突标的paper的索引
            if sorted_non_conflict:
                earliest_non_conflict_index = 0  # 因为已经按category和时间排序，第一个就是最早的
                
                # 将冲突标记paper插入到最早无冲突标的paper的正上方
                # 越晚提交的越靠近（所以在插入时，晚提交的应该先插入）
                for conflict_paper in reversed(conflict_marked_papers):  # 反向遍历，晚提交的先处理
                    sorted_non_conflict.insert(earliest_non_conflict_index, conflict_paper)
            else:
                # 如果没有无冲突标的paper，直接将冲突标记paper按提交时间倒序排列
                sorted_non_conflict = sorted(conflict_marked_papers, key=lambda x: x.submission_time, reverse=True)
        
        existing_papers = sorted_non_conflict

        # 保存更新后的数据库
        df = self.get_dataframe_from_papers(existing_papers)
        password = self.get_password()
        success = self.save_database(df, password)
        
        if success:
            return added_papers, conflict_papers
        else:
            print("保存数据库失败，添加的论文未能保存，将视为冲突paper保留在模板文件内")
            return [], new_papers  # 如果保存失败，返回空列表和所有新论文作为冲突
    
    def update_paper(self, paper: Paper, updates: Dict[str, Any]) -> bool:
        """更新单篇论文"""
        df = self.load_database()
        papers = self.get_papers_from_dataframe(df)
        success = False

        updated_papers=[]
        # 查找论文
        target_papers = [p for p in papers if is_same_identity(p, paper)]

        for paper in papers:
            if paper in target_papers:
                if success == False:
                    success = True
                    # 应用更新
                    for key, value in updates.items():
                        if hasattr(paper, key):
                                setattr(paper, key, value)
                    updated_papers.append(paper)
                else:
                    continue # 只更新第一篇找到的，其他的过滤掉
            else:
                updated_papers.append(paper)
                
        # 保存更新
        df = self.get_dataframe_from_papers(updated_papers)
        password = self.get_password()
        return self.save_database(df, password)
        
        return False
    
    def delete_paper(self, paper: Paper) -> bool:
        """删除单篇论文"""
        df = self.load_database()
        papers = self.get_papers_from_dataframe(df)
        
        # 过滤掉要删除的论文
        filtered_papers = [p for p in papers if not is_same_identity(p, paper)]

        if len(filtered_papers) < len(papers):
            # 有论文被删除
            df = self.get_dataframe_from_papers(filtered_papers)
            password = self.get_password()
            return self.save_database(df, password)
        
        return False