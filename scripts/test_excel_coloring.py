"""Test if Excel cells with invalid_fields are colored red"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from src.core.database_manager import DatabaseManager
from src.core.database_model import Paper
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def test_excel_coloring():
    """Test Excel coloring for invalid_fields"""
    
    # Create database manager
    db = DatabaseManager()
    
    # Load current database
    df = db.load_database()
    papers = db.update_utils.excel_to_paper(df, only_non_system=False, skip_invalid=False)
    
    print(f"\n=== Excel 着色验证测试 ===")
    print(f"总论文数: {len(papers)}")
    
    # Find papers with invalid_fields
    papers_with_invalid = [p for p in papers if getattr(p, 'invalid_fields', '')]
    print(f"具有 invalid_fields 的论文数: {len(papers_with_invalid)}")
    
    if papers_with_invalid:
        print(f"\n前3篇有 invalid_fields 的论文:")
        for i, p in enumerate(papers_with_invalid[:3]):
            print(f"  {i+1}. Title: {p.title[:50]}...")
            print(f"     invalid_fields: {p.invalid_fields}")
    
    # Check Excel file coloring
    excel_path = db.core_excel_path
    print(f"\n=== 检查 Excel 文件着色 ===")
    print(f"Excel 文件: {excel_path}")
    
    if os.path.exists(excel_path):
        # Load workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Get column letter for invalid_fields
        invalid_col_name = db.config.get_tag_field("invalid_fields", "table_name")
        print(f"invalid_fields 列名: {invalid_col_name}")
        
        if invalid_col_name in df.columns:
            col_idx = df.columns.get_loc(invalid_col_name)
            print(f"列索引: {col_idx}, 列字母: {chr(65 + col_idx)}")
            
            # Check cells for red color
            invalid_color = db.settings.get('excel', {}).get('invalid_fill_color', 'FF0000')
            print(f"预期无效字段填充颜色: {invalid_color}")
            
            red_count = 0
            red_rows = []
            
            # Check first 30 rows
            for row_idx in range(2, min(32, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                
                if cell.value and cell.value not in ['0', '', None, 'nan', 'false']:
                    # Check if cell has red fill
                    if cell.fill and hasattr(cell.fill, 'start_color'):
                        fill_color = str(cell.fill.start_color.rgb) if cell.fill.start_color else None
                        
                        # Check if it's red (could be FF0000 or other variations)
                        if fill_color and ('FF0000' in str(fill_color) or 'FFFF0000' in str(fill_color)):
                            red_count += 1
                            red_rows.append(row_idx)
                            print(f"✓ Row {row_idx}: Value={cell.value}, Color={fill_color}")
                        else:
                            print(f"✗ Row {row_idx}: Value={cell.value}, Color={fill_color} (不是红色)")
                    else:
                        print(f"✗ Row {row_idx}: Value={cell.value}, 没有填充颜色")
            
            print(f"\n已检查行数: 30行")
            print(f"标红单元格数: {red_count}")
            if red_rows:
                print(f"标红行号: {red_rows}")
        else:
            print(f"✗ 列 '{invalid_col_name}' 不在 dataframe 中")
            print(f"Available columns: {list(df.columns)}")
    else:
        print(f"✗ Excel 文件不存在: {excel_path}")

if __name__ == "__main__":
    test_excel_coloring()
