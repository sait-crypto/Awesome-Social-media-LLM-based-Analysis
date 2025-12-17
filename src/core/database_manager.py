"""
数据库管理器
处理Excel数据库的读写操作
"""
import os,sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Any, Tuple
import shutil
from datetime import datetime
import warnings
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../'))

# 导入配置和模型
from src.core.config_loader import get_config_instance
from src.core.database_model import Paper, is_same_identity, is_duplicate_paper
from src.core.update_file_utils import get_update_file_utils


class DatabaseManager:
    """数据库管理器"""
    
    def __init__(self):
        self.config = get_config_instance()
        self.settings = get_config_instance().settings
        self.core_excel_path = self.settings['paths']['core_excel']
        self.backup_dir = self.settings['paths']['backup_dir']
        self.conflict_marker = self.settings['database']['conflict_marker']
        
        self.update_utils = get_update_file_utils()

        # 确保目录存在
        os.makedirs(os.path.dirname(self.core_excel_path), exist_ok=True)
        os.makedirs(self.backup_dir, exist_ok=True)
    
    def backup_database(self) -> str:
        """备份数据库，返回备份文件路径"""
        if not os.path.exists(self.core_excel_path):
            return ""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"paper_database_backup_{timestamp}.xlsx"
        backup_path = os.path.join(self.backup_dir, backup_filename)
        
        shutil.copy2(self.core_excel_path, backup_path)
        print(f"数据库已备份到: {backup_path}")
        return backup_path
    
    def load_database(self) -> pd.DataFrame:
        """加载数据库到DataFrame"""
        if not os.path.exists(self.core_excel_path):
            # 创建新的数据库文件
            return self._create_new_database()
        
        try:
            # 尝试读取Excel文件
            df = pd.read_excel(self.core_excel_path, engine='openpyxl')
            
            # 确保所有必需的列都存在
            return self._ensure_columns_exist(df)
        except Exception as e:
            print(f"加载数据库失败: {e}")
            # 创建新的数据库文件
            return self._create_new_database()
    
    def _create_new_database(self) -> pd.DataFrame:
        """创建新的DataFrame结构（根据tag config的列）"""
        active_tags = self.config.get_active_tags()
        active_tags.sort(key=lambda x: x.get('order', 0))
        cols = [t['table_name'] for t in active_tags]
        df = pd.DataFrame(columns=cols)
        # 初始化时把 header style 信息存入 DataFrame.attrs，写回时会被应用到Excel
        header_fill, required_fill, required_font = self.update_utils.get_header_styles()
        df.attrs['header_styles'] = {
            'header_row_fill': header_fill,
            'required_header_fill': required_fill,
            'required_font_color': required_font
        }
        return df
    
    def _ensure_columns_exist(self, df: pd.DataFrame) -> pd.DataFrame:
        """确保DataFrame包含所有必需的列"""
        active_tags = self.config.get_active_tags()
        active_tags.sort(key=lambda x: x['order'])
        required_columns = [tag['table_name'] for tag in active_tags]
        
        # 添加缺失的列
        for column in required_columns:
            if column not in df.columns:
                df[column] = None
        
        # 移除多余的列（不在激活标签中的列）
        columns_to_keep = [col for col in df.columns if col in required_columns]
        
        # 重新排序列
        df = df[required_columns]
        
        return df
    
    def save_database(self, df: pd.DataFrame, password: str = "") -> bool:
        """保存DataFrame到Excel文件"""
        try:
            # 先根据 tag_config 规范 DataFrame 列与 category 值
            df = self.update_utils.normalize_dataframe_columns(df, self.config)

            # 备份原文件
            self.backup_database()
            
            # 保存到Excel
            with pd.ExcelWriter(
                self.core_excel_path,
                engine='openpyxl'
            ) as writer:
                df.to_excel(writer, index=False, sheet_name='Papers')
                
                # 获取工作簿和工作表
                workbook = writer.book
                worksheet = writer.sheets['Papers']
                
                # 应用格式
                self._apply_excel_formatting(workbook, worksheet, df)
                
                # 如果有密码，尝试设置保护（但openpyxl的写保护有限）
                if password:
                    try:
                        worksheet.protection.set_password(password)
                        worksheet.protection.sheet = True
                    except:
                        print("注意：无法设置Excel保护，文件将以未加密形式保存")
            
            print(f"数据库已保存到: {self.core_excel_path}")
            return True
        except Exception as e:
            print(f"保存数据库失败: {e}")
            return False
    
    
    
    def _apply_excel_formatting(self, workbook, worksheet, df):
        """对Excel应用列宽、表头格式等美化"""
 
        self.update_utils.apply_excel_formatting(workbook, worksheet, df)
        
        # 标记冲突行
        conflict_row_name=self.config.get_tag_field("conflict_marker","table_name")
        if conflict_row_name in df.columns:
            conflict_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            for idx, row in df.iterrows():
                if row.get(conflict_row_name) not in [False,0,None,"False","FALSE","false","","0"]:
                    for cell in worksheet[idx + 2]:  # +2因为标题行是1，索引从0开始
                        cell.fill = conflict_fill
                        #如果列名是doi，就在该cell内容前加冲突标记
                        if cell.value and cell.column_letter == get_column_letter(df.columns.get_loc('doi') + 1)\
                            and cell.value.startswith(self.conflict_marker)==False:
                            cell.value = f"{self.conflict_marker} {cell.value}"
        else:
            print(f"添加论文冲突格式时，发现数据库表格没有conflict_marker列")

                        
    
    def get_password(self) -> str:
        """获取Excel密码"""
        # 首先尝试从环境变量获取
        password = os.environ.get('EXCEL_KEY', '')
        if password:
            return password
        
        # 尝试从本地文件获取
        password_path = self.settings['excel'].get('password_path', '')
        if password_path and os.path.exists(password_path):
            try:
                with open(password_path, 'r', encoding='utf-8') as f:
                    password = f.read().strip()
                    return password
            except Exception as e:
                print(f"本地读取Excel密码失败{e}")
                
        
        return ""  # 返回空字符串表示不加密
    
    
    
    #唯一对外接口
    def add_papers(self, new_papers: List[Paper], conflict_resolution: str = 'mark') -> Tuple[List[Paper], List[Paper]]:
        """
        添加新论文到数据库，同时验证冲突
        
        参数:
            new_papers: 新论文列表
            conflict_resolution: 冲突解决策略 ('mark', 'skip', 'replace')
        
        返回:
            Tuple[成功添加的论文列表, 冲突论文列表（被标记后需要加入数据库的）]
        """
        # 加载现有数据库
        df = self.load_database()
        old_papers = self.update_utils.excel_to_paper(df, only_non_system=False)

        # non_conflict_papers存储结构: [(主论文, [冲突论文1, 冲突论文2, ...]), ...]
        non_conflict_papers: List[Tuple[Paper, List[Paper]]] = []    
        conflict_papers: List[Paper] = []
        added_papers = []
        old_conflict_papers = []
        
        # 分离现有论文中的正常论文和冲突论文
        for p in old_papers:
            if not p.conflict_marker:
                non_conflict_papers.append((p, []))
            else:
                old_conflict_papers.append(p)
        
        # 还原原有的冲突结构
        for old_conflict in old_conflict_papers:
            conflict_found = False
            for i, (main_paper, conflict_list) in enumerate(non_conflict_papers):
                if is_same_identity(old_conflict, main_paper):
                    conflict_list.append(old_conflict)
                    conflict_found = True
                    break
            
            if not conflict_found:
                # 如果没有找到对应的主论文，将这个冲突论文作为一个新的主论文
                print(f"警告：数据库原有冲突论文 {old_conflict.title[:50]}... 没有找到对应主论文，将其作为新论文添加")
                non_conflict_papers.append((old_conflict, []))
                old_conflict.conflict_marker = False  # 清除冲突标记
        
        # 处理新论文
        for new_paper in new_papers:
            # 检查是否已存在
            same_identity_indices = []
            main_paper_idx = -1
            
            # 找出所有同identity论文
            for idx, (main_paper, conflict_list) in enumerate(non_conflict_papers):
                if is_same_identity(new_paper, main_paper):
                    same_identity_indices.append(idx)
                    if not main_paper.conflict_marker and main_paper_idx == -1:
                        main_paper_idx = idx
            
            if same_identity_indices:
                # 检查是否为"完全重复提交" - 需要包括主论文和所有冲突论文
                all_same_papers = []
                for idx in same_identity_indices:
                    main_paper, conflict_list = non_conflict_papers[idx]
                    all_same_papers.append(main_paper)
                    all_same_papers.extend(conflict_list)
                
                if is_duplicate_paper(all_same_papers, new_paper, complete_compare=False):
                    print(f"论文: {new_paper.title}——重复提交，跳过添加")
                    continue
                
                # 不是完全相同，按冲突策略处理
                if conflict_resolution == 'skip':
                    print(f"论文: {new_paper.title}——存在冲突，已跳过")
                    continue
                
                elif conflict_resolution == 'replace':
                    # 完全替换：删除所有同身份论文，添加新论文
                    # 从后往前删除，避免索引错乱
                    for idx in sorted(same_identity_indices, reverse=True):
                        del non_conflict_papers[idx]
                    
                    non_conflict_papers.append((new_paper, []))
                    added_papers.append(new_paper)
                    print(f"论文: {new_paper.title}——替换原有论文")
                
                elif conflict_resolution == 'mark':
                    new_paper.conflict_marker = True
                    
                    if main_paper_idx != -1:
                        # 添加到对应主论文的冲突列表中
                        non_conflict_papers[main_paper_idx][1].append(new_paper)
                        conflict_papers.append(new_paper)
                        print(f"论文: {new_paper.title}——与现有论文冲突，已标记并作为冲突论文添加")
                    else:
                        # 所有同身份论文都有冲突标记，这是一个特殊情况
                        # 将第一篇作为主论文，并添加冲突
                        first_idx = same_identity_indices[0]
                        main_paper = non_conflict_papers[first_idx][0]
                        conflict_list = non_conflict_papers[first_idx][1]
                        
                        # 清除第一篇的冲突标记，使其成为主论文
                        main_paper.conflict_marker = False
                        conflict_list.append(new_paper)
                        conflict_papers.append(new_paper)
                        print(f"警告：论文 {new_paper.title[:50]}... 的所有同身份论文都有冲突标记，将第一篇清除标记并作为主论文")
            else:
                # 新论文，添加到列表
                non_conflict_papers.append((new_paper, []))
                added_papers.append(new_paper)
                print(f"论文: {new_paper.title}——作为新论文添加")
        
        # 排序
        # 按category分组
        category_groups = {}
        for main_paper, conflict_list in non_conflict_papers:
            if main_paper.category not in category_groups:
                category_groups[main_paper.category] = []
            category_groups[main_paper.category].append((main_paper, conflict_list))
        
        # 每个category内部按提交时间倒序（假设有submission_time属性）
        sorted_all_papers = []
        for category in sorted(category_groups.keys()):
            papers_in_category = category_groups[category]
            # 按主论文的提交时间倒序排序
            papers_in_category.sort(key=lambda x: x[0].submission_time, reverse=True)
            
            # 添加主论文和对应的冲突论文
            for main_paper, conflict_list in papers_in_category:
                # 先添加冲突论文（按提交时间倒序，最新的在最上面）
                if conflict_list:
                    conflict_list.sort(key=lambda x: x.submission_time, reverse=True)
                    sorted_all_papers.extend(conflict_list)
                
                # 再添加主论文
                sorted_all_papers.append(main_paper)
        
        # 保存更新后的数据库
        df = self.update_utils.paper_to_excel(sorted_all_papers, only_non_system=False)
        password = self.get_password()
        success = self.save_database(df, password)
        
        if success:
            return added_papers, conflict_papers
        else:
            print("保存数据库失败，添加的论文未能保存")
            return [], new_papers
    
    def update_paper(self, paper: Paper, updates: Dict[str, Any]) -> bool:
        """更新单篇论文"""
        df = self.load_database()
        papers = self.update_utils.excel_to_paper(df,only_non_system=False)
        success = False

        updated_papers=[]
        # 查找论文
        target_papers = [p for p in papers if is_same_identity(p, paper)]

        for paper in papers:
            if paper in target_papers:
                if success == False:
                    success = True
                    # 应用更新
                    for key, value in updates.items():
                        if hasattr(paper, key):
                                setattr(paper, key, value)
                    updated_papers.append(paper)
                else:
                    continue # 只更新第一篇找到的，其他的过滤掉
            else:
                updated_papers.append(paper)
                
        # 保存更新
        df = self.update_utils.paper_to_excel(updated_papers,only_non_system=False) 
        password = self.get_password()
        return self.save_database(df, password)
        
    
    def delete_paper(self, paper: Paper) -> bool:
        """删除单篇论文"""
        df = self.load_database()
        papers = self.update_utils.excel_to_paper(df,only_non_system=False)
        
        # 过滤掉要删除的论文
        filtered_papers = [p for p in papers if not is_same_identity(p, paper)]

        if len(filtered_papers) < len(papers):
            # 有论文被删除
            df = self.update_utils.paper_to_excel(filtered_papers,only_non_system=False)
            password = self.get_password()
            return self.save_database(df, password)
        
        return False