"""
更新文件工具模块
统一处理模板文件（Excel和JSON）的读取、写入和移除操作
对于json操作，保证不应使用任何非基础第三方包，以供submit_gui调用
提供以下方法：
read_json_file
write_json_file
read_excel_file
write_excel_file
load_papers_from_excel
load_papers_from_json
save_papers_to_json  <-- 新增
remove_papers_from_json
remove_papers_from_excel
persist_ai_generated_to_update_files
normalize_update_file_columns
create_empty_update_file_df
normalize_category_value
normalize_dataframe_columns
excel_to_paper
json_to_paper
paper_to_excel
paper_to_json
"""
import os
import json
from typing import List, Dict, Any, Optional,Union,Tuple
from dataclasses import asdict

from src.core.config_loader import get_config_instance
from src.core.database_model import Paper,is_same_identity, is_duplicate_paper
# 导入统一的备份函数
from src.utils import ensure_directory, backup_file, get_current_timestamp

class UpdateFileUtils:
    """更新文件工具类"""
    
    def __init__(self):
        self.config = get_config_instance()
        self.settings = get_config_instance().settings
        self.update_excel_path = self.settings['paths']['update_excel']
        self.update_json_path = self.settings['paths']['update_json']
        self.backup_dir = self.settings['paths']['backup_dir']


    def read_json_file(self,filepath: str) -> Optional[Dict]:
        """读取JSON文件"""
        try:
            if not os.path.exists(filepath):
                return None
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"读取JSON文件失败 {filepath}: {e}")
            return None



    def write_json_file(self,filepath: str, data: Dict, indent: int = 2) -> bool:
        """写入JSON文件"""
        try:
            ensure_directory(os.path.dirname(filepath))
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=indent)
            return True
        except Exception as e:
            print(f"写入JSON文件失败 {filepath}: {e}")
            return False


    def read_excel_file(self,filepath: str) :
        """读取Excel文件"""
        try:
            import pandas as pd

        except Exception as e:
            print( f"无法导入pandas依赖:{e}\n 注意如果要加载excel文件，你需要安装pandas依赖包")
            return
        try:
            if not os.path.exists(filepath):
                return pd.DataFrame()
            
            df = pd.read_excel(filepath, engine='openpyxl')
            return df
        except Exception as e:
            print(f"读取Excel文件失败 {filepath}: {e}")
            return None


    def write_excel_file(self,filepath: str, df) -> bool:
        """写入Excel文件"""
        try:
            import pandas as pd

        except Exception as e:
            print( f"无法导入pandas依赖:{e}\n 注意如果要加载excel文件，你需要安装pandas依赖包")
            return False
        try:
            ensure_directory(os.path.dirname(filepath))
            with pd.ExcelWriter(
                filepath,
                engine='openpyxl'
            ) as writer:
                df.to_excel(writer, index=False, sheet_name='Papers')
                workbook = writer.book
                worksheet = writer.sheets['Papers']
                # 应用格式
                self.apply_excel_formatting(workbook, worksheet, df)
                return True
        except Exception as e:
            print(f"写入Excel文件失败 {filepath}: {e}")
            return False
        

    
    def load_papers_from_excel(self, filepath: str = None) -> List[Paper]:
        """从Excel文件加载论文"""

        if filepath is None:
            filepath = self.update_excel_path
            
        df = self.read_excel_file(filepath)
        if df is None or df.empty:
            return []
        
        return self.excel_to_paper(df, only_non_system=True)
    
    def load_papers_from_json(self, filepath: str = None) -> List[Paper]:
        """从JSON文件加载论文 (支持 {'papers': [...], 'meta': ...} 结构)"""
        if filepath is None:
            filepath = self.update_json_path
            
        data = self.read_json_file(filepath)
        if not data:
            return []
        
        # 提取论文列表数据
        papers_data = []
        if isinstance(data, dict):
            # 新结构 or 旧结构(单对象)
            if 'papers' in data and isinstance(data['papers'], list):
                papers_data = data['papers']
            else:
                # 可能是单个论文对象或旧结构
                # 检查是否看起来像论文数据（有doi或title）
                if 'doi' in data or 'title' in data:
                    papers_data = [data]
                else:
                    # 可能是空的 {'meta':...}
                    papers_data = []
        elif isinstance(data, list):
            # 旧结构(列表)
            papers_data = data
        
        return self.json_to_paper(papers_data, only_non_system=True)
    
    def save_papers_to_json(self, filepath: str, papers: List[Paper]) -> bool:
        """
        保存论文列表到JSON文件，自动处理 meta 结构
        结构: {'papers': [...], 'meta': {'generated_at': ...}}
        """
        try:
            # 1. 尝试读取现有文件的 meta 信息以保留其他字段
            existing_data = self.read_json_file(filepath)
            meta = {}
            if existing_data and isinstance(existing_data, dict) and 'meta' in existing_data:
                meta = existing_data['meta']
            
            # 2. 更新 generated_at
            meta['generated_at'] = get_current_timestamp()

            # 3. 转换论文数据
            papers_data = self.paper_to_json(papers)
            if isinstance(papers_data, dict): # 如果只有一条数据paper_to_json可能返回dict，强制转list
                papers_data = [papers_data]

            # 4. 构建完整数据
            final_data = {
                "papers": papers_data,
                "meta": meta
            }

            # 5. 写入
            return self.write_json_file(filepath, final_data)

        except Exception as e:
            print(f"保存论文到JSON失败 {filepath}: {e}")
            raise e

    def remove_papers_from_json(self, processed_papers: List[Paper], filepath: str = None):
        """从JSON文件中移除已处理的论文"""
        if filepath is None:
            filepath = self.update_json_path
            
        data = self.read_json_file(filepath)
        if not data:
            return
        
        # 1. 准备 Key 集合
        processed_dois = set()
        processed_titles = set()
        for paper in processed_papers:
            doi, title = paper.get_key()
            if doi: processed_dois.add(doi)
            if title: processed_titles.add(title)
        
        def should_remove(item: Dict) -> bool:
            raw_doi = str(item.get('doi', '')).strip()
            from src.utils import validate_doi
            _, n_doi = validate_doi(raw_doi, check_format=False)
            item_doi = n_doi.lower()
            
            item_title = str(item.get('title', '')).strip().lower()
            
            if item_doi and item_doi in processed_dois:
                return True
            if item_title and item_title in processed_titles:
                return True
            return False
        
        # 2. 过滤数据
        has_changes = False
        new_data = None

        if isinstance(data, list):
            original_len = len(data)
            filtered_data = [item for item in data if not should_remove(item)]
            if len(filtered_data) < original_len:
                has_changes = True
                new_data = filtered_data
            else:
                new_data = data

        elif isinstance(data, dict) and 'papers' in data:
            if isinstance(data['papers'], list):
                original_len = len(data['papers'])
                filtered_papers = [item for item in data['papers'] if not should_remove(item)]
                if len(filtered_papers) < original_len:
                    has_changes = True
                    data['papers'] = filtered_papers
                    new_data = data
                else:
                    new_data = data

        # 3. 备份并写入
        if has_changes:
            # 调用统一备份函数
            backup_file(filepath, self.backup_dir)
            
            if isinstance(new_data, list):
                self.write_json_file(filepath, new_data)
            else:
                self.write_json_file(filepath, new_data)

    def remove_papers_from_excel(self, processed_papers: List[Paper], filepath: str = None):
        """从Excel文件中移除已处理的论文"""
        try:
            import pandas as pd
        except Exception as e:
            print(f"无法导入pandas依赖:{e}")
            return
            
        if filepath is None:
            filepath = self.update_excel_path
            
        df = self.read_excel_file(filepath)
        if df is None or df.empty:
            return
        
        # 1. 准备 Key 集合
        processed_dois = set()
        processed_titles = set()
        for paper in processed_papers:
            doi, title = paper.get_key()
            if doi: processed_dois.add(doi)
            if title: processed_titles.add(title)
        
        rows_to_keep = []
        rows_removed_count = 0

        # 2. 过滤数据
        for idx, row in df.iterrows():
            raw_doi = str(row.get('doi', '')) if pd.notna(row.get('doi')) else ''
            from src.utils import validate_doi
            _, n_doi = validate_doi(raw_doi.strip(), check_format=False)
            row_doi = n_doi.lower()
            
            raw_title = str(row.get('title', '')) if pd.notna(row.get('title')) else ''
            row_title = raw_title.strip().lower()
            
            is_processed = False
            if row_doi and row_doi in processed_dois:
                is_processed = True
            elif row_title and row_title in processed_titles:
                is_processed = True
            
            if not is_processed:
                rows_to_keep.append(row)
            else:
                rows_removed_count += 1
        
        # 3. 备份并写入
        if rows_removed_count > 0:
            # 调用统一备份函数
            backup_file(filepath, self.backup_dir)
            
            if rows_to_keep:
                new_df = pd.DataFrame(rows_to_keep)
                new_df = self.normalize_update_file_columns(new_df)
                self.write_excel_file(filepath, new_df)
            else:
                empty_df = self.create_empty_update_file_df()
                self.write_excel_file(filepath, empty_df)
    
    def persist_ai_generated_to_update_files(self, papers: List[Paper], file_path: str):
        """
        把 AI 生成的字段原样写回到指定文件。
        只写回非系统字段。
        匹配策略：优先按 DOI 匹配，若 DOI 不存在则按 title（忽略大小写、首尾空白）匹配。
        **修改：根据 file_path 的扩展名决定调用 JSON 还是 Excel 处理方法。**
        """
        if not papers:
            return

        # 获取AI相关字段（都是非系统字段）
        ai_fields = ['title_translation', 'analogy_summary',
                    'summary_motivation', 'summary_innovation',
                    'summary_method', 'summary_conclusion', 'summary_limitation']
        
        # 扩展名检查
        ext = os.path.splitext(file_path)[1].lower()
        
        try:
            if ext == '.json':
                # ---------- JSON 处理 ----------
                self._persist_ai_to_json(papers, ai_fields, file_path)
            elif ext in ['.xlsx', '.xls']:
                # ---------- Excel 处理 ----------
                self._persist_ai_to_excel(papers, ai_fields, file_path)
            else:
                print(f"警告: 不支持的文件类型用于回写AI内容: {file_path}")
                return
            
            print(f"已将AI生成内容回写到更新文件: {file_path}")
        except Exception as e:
            err = f"回写AI生成内容到更新文件 {file_path} 失败: {e}"
            print(err)
            raise
    
    def _persist_ai_to_json(self, papers: List[Paper], ai_fields: List[str], file_path: str):
        try:
            
            # 读取现有数据并转换为Paper对象列表
            json_data = self.read_json_file(file_path) or {}
            
            if 'papers' not in json_data:
                json_data['papers'] = []
            
            existing_papers = self.json_to_paper(json_data['papers'], only_non_system=True)
            
            # 匹配并更新AI字段
            for ai_paper in papers:
                for existing_paper in existing_papers:
                    if is_same_identity(ai_paper, existing_paper):
                        # 更新AI相关字段
                        for field in ai_fields:
                            new_value = getattr(ai_paper, field, "")
                            if new_value:
                                setattr(existing_paper, field, new_value)
                        break
            
            # 转换回JSON格式并保存，注意保留Meta
            self.save_papers_to_json(file_path, existing_papers)
        
        except Exception as e:
            raise RuntimeError(f"写入更新JSON失败: {e}")
    
    def _persist_ai_to_excel(self, papers: List[Paper], ai_fields: List[str], file_path: str):
        try:            
            # 读取现有数据
            df = self.read_excel_file(file_path)
            if df is None or df.empty:
                return
            
            # 转换为Paper对象列表
            existing_papers = self.excel_to_paper(df, only_non_system=True)
            
            # 匹配并更新AI字段
            for ai_paper in papers:
                for existing_paper in existing_papers:
                    if is_same_identity(ai_paper, existing_paper):
                        # 更新AI相关字段
                        for field in ai_fields:
                            new_value = getattr(ai_paper, field, "")
                            if new_value:
                                setattr(existing_paper, field, new_value)
                        break
            
            # 转换回DataFrame并保存
            new_df = self.paper_to_excel(existing_papers, only_non_system=True)
            self.write_excel_file(file_path, new_df)
            
        except Exception as e:
            raise RuntimeError(f"写入更新Excel失败: {e}")
    

    
    def normalize_update_file_columns(self, df):
        """
        确保DataFrame列按照非系统字段重新生成
        """
        try:
            import pandas as pd

        except Exception as e:
            print( f"无法导入pandas依赖:{e}\n 注意如果要加载excel文件，你需要安装pandas依赖包")
            return
        # 只使用非系统字段
        non_system_tags = self.config.get_non_system_tags()
        non_system_tags.sort(key=lambda x: x['order'])
        columns = [tag['table_name'] for tag in non_system_tags]
        
        # 添加缺失列
        for c in columns:
            if c not in df.columns:
                df[c] = ""
        
        # 重新排序和保留需要的列
        df = df.reindex(columns=columns)
        
        # 类型转换
        for tag in non_system_tags:
            col = tag['table_name']
            t = tag.get('type', 'string')
            if col in df.columns:
                if t == 'bool':
                    df[col] = df[col].fillna(False).astype(bool)
                elif t == 'int':
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
                else:
                    df[col] = df[col].fillna("").astype(str).str.strip()
        # 保存 header style 到 attrs，确保写回Excel时能恢复表头颜色
        header_fill, required_fill, required_font = self.get_header_styles()
        df.attrs['header_styles'] = {
            'header_row_fill': header_fill,
            'required_header_fill': required_fill,
            'required_font_color': required_font
        }
        return df
    
    def create_empty_update_file_df(self):
        """创建空的更新文件DataFrame（只包含非系统字段）"""
        try:
            import pandas as pd

        except Exception as e:
            print( f"无法导入pandas依赖:{e}\n 注意如果要加载excel文件，你需要安装pandas依赖包")
            return
        non_system_tags = self.config.get_non_system_tags()
        non_system_tags.sort(key=lambda x: x['order'])
        columns = [tag['table_name'] for tag in non_system_tags]
        df= pd.DataFrame(columns=columns)
        # 把 header style 信息存入 DataFrame.attrs，以便写回Excel时保留表头样式
        header_fill, required_fill, required_font = self.get_header_styles()
        df.attrs['header_styles'] = {
            'header_row_fill': header_fill,
            'required_header_fill': required_fill,
            'required_font_color': required_font
        }

        return df
    

    def ensure_update_file_format(self, filepath: str = None) -> bool:
        """确保更新文件按非系统字段规范化并写回，保证表头样式存在。
        返回True表示写回成功或文件存在并已格式化，False表示失败。
        """
        try:
            import pandas as pd
        except Exception as e:
            print( f"无法导入pandas依赖:{e}\n 注意如果要加载excel文件，你需要安装pandas依赖包")
            return False

        if filepath is None:
            # 默认只处理基础的 excel 路径，如果需要处理其他路径，调用者应传入 filepath
            filepath = self.update_excel_path

        # 检查文件扩展名，只处理Excel
        if not (filepath.endswith('.xlsx') or filepath.endswith('.xls')):
            return False

        df = self.read_excel_file(filepath)
        if df is None or df.empty:
            df = self.create_empty_update_file_df()
        else:
            try:
                df = self.normalize_update_file_columns(df)
            except Exception as e:
                print(f"规范化更新文件列失败: {e}")
                return False

        header_fill, required_fill, required_font = self.get_header_styles()
        df.attrs['header_styles'] = {
            'header_row_fill': header_fill,
            'required_header_fill': required_fill,
            'required_font_color': required_font
        }

        return self.write_excel_file(filepath, df)

    def _regenerate_columns_from_tags(self,config_instance) -> List[str]:
        """根据tag_config生成按order排序的表列名（table_name）列表"""
        active_tags = config_instance.get_active_tags()
        active_tags.sort(key=lambda x: x.get('order', 0))
        return [tag['table_name'] for tag in active_tags]

    #暂不实现，因为需要将旧name保存下来以供映射
    def normalize_category_value(self,raw_val: Any, config_instance) -> str:
        """
        把 category 字段的旧name改为新的name，
        若无法匹配则返回原值的字符串形式（strip 后）。
        """
        return raw_val
        if raw_val is None:
            return ""
        val = str(raw_val).strip()
        if not val:
            return ""
        # 构建映射：旧name -> unique_name, unique_name ->新name
        new_cats = config_instance.get_active_categories()
        old_cats=config_instance.old_cats


    def normalize_dataframe_columns(self,df, config_instance) -> Any:
        """
        确保DataFrame列按照tag_config中active tags重新生成：
        - 添加缺失列（置空）
        - 移除未激活的列
        - 按order排序列顺序
        - 对 category 列内值执行规范化（未实现）
        """
        try:
            import pandas as pd

        except Exception as e:
            print( f"无法导入pandas依赖:{e}\n 注意如果要加载excel文件，你需要安装pandas依赖包")
            return 
        if df is None:
            df = pd.DataFrame()
        cols = self._regenerate_columns_from_tags(config_instance)
        # 保留现有行数据但只保留激活列
        for c in cols:
            if c not in df.columns:
                df[c] = ""
        # 移除多余列
        to_keep = cols
        df = df.loc[:, [c for c in to_keep if c in df.columns]]
        # reorder
        df = df[cols]
        # # 规范化 category 列值
        # if 'category' in df.columns:
        #     df['category'] = df['category'].apply(lambda v: normalize_category_value(v, config_instance))
        # 将所有非-bool/int 列转为 string（保持原有语义）
        for tag in config_instance.get_active_tags():
            col = tag['table_name']
            t = tag.get('type', 'string')
            if col in df.columns:
                if t == 'bool':
                    df[col] = df[col].fillna(False).astype(bool)
                elif t == 'int':
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
                else:
                    df[col] = df[col].fillna("").astype(str).str.strip()
        # 保存 header style 到 attrs，确保写回Excel时能恢复表头颜色
        header_fill, required_fill, required_font = self.get_header_styles()
        df.attrs['header_styles'] = {
            'header_row_fill': header_fill,
            'required_header_fill': required_fill,
            'required_font_color': required_font
        }
        return df

    def normalize_json_papers(self,raw_papers: List[Dict[str, Any]], config_instance) -> List[Dict[str, Any]]:
        """
        把JSON中的每篇论文都规范化为只包含active tag的变量（使用variable作为键），
        并将类型与category规范化。（未实现）
        """
        normalized_list = []
        active_tags = config_instance.get_active_tags()
        for item in raw_papers:
            out = {}
            for tag in active_tags:
                var = tag['variable']
                table_name = tag['table_name']
                # 支持输入既有 variable 也有 table_name 两种键
                val = item.get(var, item.get(table_name, ""))
                if val is None:
                    val = ""
                t = tag.get('type', 'string')
                if t == 'bool':
                    out[var] = bool(val) if val not in ("", None) else False
                elif t == 'int':
                    try:
                        out[var] = int(val)
                    except Exception:
                        out[var] = 0
                else:
                    out[var] = str(val).strip()
            # 规范化 category 存储为 unique_name
            # if 'category' in out:
            #     out['category'] = normalize_category_value(out.get('category', ""), config_instance)
            normalized_list.append(out)
        return normalized_list
    
    #===================数据规范化===========================
    def json_to_paper(self, json_data: Union[Dict, List[Dict]], only_non_system: bool = False) -> List[Paper]:
        """
        数据规范化方法：将JSON数据转换为Paper对象列表
        """
        # 统一处理为列表
        if isinstance(json_data, dict):
            data_list = [json_data]
        else:
            data_list = json_data
        
        # 获取标签配置
        if only_non_system:
            tags = self.config.get_non_system_tags()
        else:
            tags = self.config.get_active_tags()
        
        # 规范化JSON数据
        normalized_list = self.normalize_json_papers(data_list, self.config)
        
        papers = []
        for item in normalized_list:
            try:
                # 转换为Paper对象
                paper_data = self._dict_to_paper_data(item, tags)
                paper = Paper.from_dict(paper_data)
                
                # 验证论文字段
                valid, errors = paper.validate_paper_fields(
                    self.config,
                    check_required=False,  # 更新文件可能不完整
                    check_non_empty=True
                )
                
                if not valid:
                    print(f"警告: 跳过验证失败的论文: {paper.title[:50]}...")
                    continue
                
                papers.append(paper)
            except Exception as e:
                print(f"警告: 解析JSON条目失败: {e}")
                continue
        
        return papers
    
    def excel_to_paper(self, df, only_non_system: bool = False) -> List[Paper]:
        """
        数据规范化方法：将Excel数据转换为Paper对象列表
        """
        try:
            import pandas as pd
        except ImportError:
            print("警告: pandas未安装，无法处理Excel文件")
            return []
        
        # 统一处理为DataFrame
        if isinstance(df, pd.Series):
            df = pd.DataFrame([df])
        
        if df is None or df.empty:
            return []
        
        # 获取标签配置
        if only_non_system:
            tags = self.config.get_non_system_tags()
        else:
            tags = self.config.get_active_tags()
        
        papers = []
        for _, row in df.iterrows():
            try:
                # 转换为Paper对象
                paper_data = self._excel_row_to_paper_data(row, tags)
                paper = Paper.from_dict(paper_data)
                
                # 验证论文字段
                valid, errors = paper.validate_paper_fields(
                    self.config,
                    check_required=False,  # Excel文件可能不完整
                    check_non_empty=True
                )
                
                if not valid:
                    print(f"警告: 跳过验证失败的论文: {paper.title[:50]}...")
                    continue
                
                papers.append(paper)
            except Exception as e:
                print(f"警告: 解析Excel行失败: {e}")
                continue
        
        return papers
    
    def paper_to_json(self, papers: Union[Paper, List[Paper]]) -> Union[Dict, List[Dict]]:
        """
        数据规范化方法：将Paper对象（或列表）转换为JSON可序列化的字典（或字典列表）
        
        Args:
            papers: Paper对象或Paper对象列表
            
        Returns:
            如果是单个Paper返回字典，如果是列表返回字典列表
        """
        # 统一处理为列表
        if isinstance(papers, Paper):
            input_is_single = True
            papers_list = [papers]
        else:
            input_is_single = False
            papers_list = papers
        

        result = []
        for paper in papers_list:
            try:
                paper_dict = self._paper_to_dict(paper)
                result.append(paper_dict)
            except Exception as e:
                print(f"警告: 转换Paper到字典失败: {e}")
                continue
        paper_dict=self.normalize_json_papers(result, self.config)
        # 返回与输入一致的格式
        if input_is_single and result:
            return result[0]
        return result
    
    def paper_to_excel(self, papers: Union[Paper, List[Paper]], only_non_system: bool = False) :
        """
        数据规范化方法：将Paper对象（或列表）转换为Excel DataFrame
        
        Args:
            papers: Paper对象或Paper对象列表
            only_non_system: 是否只包含非系统字段
            
        Returns:
            DataFrame
        """
        try:
            import pandas as pd

        except Exception as e:
            print( f"无法导入pandas依赖:{e}\n 注意如果要加载excel文件，你需要安装pandas依赖包")
            return
        # 统一处理为列表
        if isinstance(papers, Paper):
            papers_list = [papers]
        else:
            papers_list = papers
        
        if not papers_list:
            return pd.DataFrame()
        
        # 获取标签配置
        if only_non_system:
            tags = self.config.get_non_system_tags()
        else:
            tags = self.config.get_active_tags()
        
        # 按order排序
        tags.sort(key=lambda x: x['order'])
        
        # 准备数据
        data = []
        for paper in papers_list:
            try:
                # 新增检查：确保paper是Paper实例
                if not isinstance(paper, Paper):
                    print(f"警告: 跳过非Paper实例: {type(paper)}")
                    continue
                row_data = self._paper_to_excel_row(paper, tags)
                data.append(row_data)
            except Exception as e:
                print(f"警告: 转换Paper到Excel行失败: {e}")
                continue
        
        # 创建DataFrame
        if not data:
            return pd.DataFrame()
        
        # 获取列名（使用table_name）
        columns = [tag['table_name'] for tag in tags]
        df = pd.DataFrame(data, columns=columns)
        df = self.normalize_dataframe_columns(df, self.config)

        
        return df
    
    # ========== 核心私有方法 ==========
    
    def _dict_to_paper_data(self, data_dict: Dict, tags: List[Dict]) -> Dict:
        """
        核心方法：将字典数据转换为Paper可用的数据字典
        
        Args:
            data_dict: 原始数据字典
            tags: 标签配置列表
            
        Returns:
            规范化的Paper数据字典
        """
        paper_data = {}
        
        for tag in tags:
            var_name = tag['variable']
            table_name = tag['table_name']
            tag_type = tag.get('type', 'string')
            
            # 尝试从两种键名获取值
            value = data_dict.get(var_name, data_dict.get(table_name, ""))
            
            # 处理空值
            if value is None or (isinstance(value, str) and value.strip() == ""):
                # 根据类型设置默认值
                if tag_type == 'bool':
                    value = False
                elif tag_type == 'int':
                    value = 0
                elif tag_type == 'float':
                    value = 0.0
                else:
                    value = ""
            else:
                # 类型转换
                value = self._convert_value_by_type(value, tag_type)
                
                # 特殊处理：验证pipeline_image字段（支持多图）
                if var_name == 'pipeline_image' and value:
                    # 验证图片格式和路径（允许多图）
                    from src.utils import validate_pipeline_image
                    fig_dir = self.config.settings['paths'].get('figure_dir', 'figures')
                    valid, normalized = validate_pipeline_image(value, fig_dir)
                    if not valid:
                        print(f"警告: 图片路径格式无效: {value}")
                        value = ""  # 如果无效，清空
                    else:
                        value = normalized
            
            paper_data[var_name] = value
        
        return paper_data
    
    def _excel_row_to_paper_data(self, row, tags: List[Dict]) -> Dict:
        """
        核心方法：将Excel行转换为Paper可用的数据字典
        
        Args:
            row: pandas Series
            tags: 标签配置列表
            
        Returns:
            规范化的Paper数据字典
        """
        try:
            import pandas as pd

        except Exception as e:
            print( f"无法导入pandas依赖:{e}\n 注意如果要加载excel文件，你需要安装pandas依赖包")
            return {}
        paper_data = {}
        
        for tag in tags:
            var_name = tag['variable']
            table_name = tag['table_name']
            tag_type = tag.get('type', 'string')
            
            # Excel中使用table_name作为列名
            if table_name in row:
                value = row[table_name]
            else:
                value = ""
            
            # 处理pandas NaN
            if pd.isna(value):
                # 根据类型设置默认值
                if tag_type == 'bool':
                    value = False
                elif tag_type == 'int':
                    value = 0
                elif tag_type == 'float':
                    value = 0.0
                else:
                    value = ""
            else:
                # 类型转换
                value = self._convert_value_by_type(value, tag_type)
            
            paper_data[var_name] = value
        
        return paper_data
    
    def _paper_to_dict(self, paper: Paper) -> Dict:
        """
        核心方法：将Paper对象转换为字典
        
        Args:
            paper: Paper对象
            
        Returns:
            字典，键为variable名称
        """
        paper_dict = asdict(paper)
        
        # 确保所有字段都是可序列化的
        for key, value in paper_dict.items():
            if value is None:
                paper_dict[key] = ""
            elif isinstance(value, bool):
                paper_dict[key] = value
            elif isinstance(value, (int, float)):
                paper_dict[key] = value
            else:
                paper_dict[key] = str(value)
        
        return paper_dict
    
    def _paper_to_excel_row(self, paper: Paper, tags: List[Dict]) -> Dict:
        """
        核心方法：将Paper对象转换为Excel行数据
        
        Args:
            paper: Paper对象
            tags: 标签配置列表
            
        Returns:
            字典，键为table_name
        """
        row_data = {}
        paper_dict = asdict(paper)
        
        for tag in tags:
            var_name = tag['variable']
            table_name = tag['table_name']
            tag_type = tag.get('type', 'string')
            
            value = paper_dict.get(var_name, "")
            
            # 处理空值
            if value is None or (isinstance(value, str) and value.strip() == ""):
                value = ""
            
            # 类型转换确保Excel中的格式正确
            if tag_type == 'bool':
                # Excel中布尔值通常显示为TRUE/FALSE
                value = bool(value)
            elif tag_type == 'int':
                try:
                    value = int(value) if value not in ("", None) else 0
                except (ValueError, TypeError):
                    value = 0
            elif tag_type == 'float':
                try:
                    value = float(value) if value not in ("", None) else 0.0
                except (ValueError, TypeError):
                    value = 0.0
            
            row_data[table_name] = value
        
        return row_data
    
    def _convert_value_by_type(self, value: Any, tag_type: str) -> Any:
        """
        根据标签类型转换值
        
        Args:
            value: 原始值
            tag_type: 标签类型
            
        Returns:
            转换后的值
        """
        if tag_type == 'bool':
            # 布尔类型转换
            if isinstance(value, bool):
                return value
            elif isinstance(value, (int, float)):
                return bool(value)
            elif isinstance(value, str):
                value_lower = value.strip().lower()
                if value_lower in ('True','TRUE','true', 'yes', '1', 'y', '是', '真'):
                    return True
                elif value_lower in ('False''FALSE','false', 'no', '0', 'n', '否', '假'):
                    return False
                else:
                    # 尝试转换为布尔值
                    try:
                        return bool(int(value))
                    except:
                        return False
            else:
                return bool(value)
        
        elif tag_type == 'int':
            # 整数类型转换
            if isinstance(value, (int, float)):
                return int(value)
            elif isinstance(value, str) and value.strip():
                try:
                    return int(float(value.strip()))
                except (ValueError, TypeError):
                    return 0
            else:
                return 0
        
        elif tag_type == 'float':
            # 浮点数类型转换
            if isinstance(value, (int, float)):
                return float(value)
            elif isinstance(value, str) and value.strip():
                try:
                    return float(value.strip())
                except (ValueError, TypeError):
                    return 0.0
            else:
                return 0.0
        
        elif tag_type == 'text':
            # 文本类型，保持原样
            if isinstance(value, (list, dict)):
                # 复杂结构转换为JSON字符串
                try:
                    return json.dumps(value, ensure_ascii=False)
                except:
                    return str(value)
            else:
                return str(value).strip()
        
        else:  # 'string' 或 'enum' 或其他
            # 字符串类型
            if isinstance(value, (list, dict)):
                # 复杂结构转换为JSON字符串
                try:
                    return json.dumps(value, ensure_ascii=False)
                except:
                    return str(value)
            else:
                return str(value).strip()
    def get_header_styles(self) -> Tuple[str, str, str]:
        """
        从 settings 中读取颜色配置（优先支持多种可能的key名称以兼容你的 setting_config）。
        返回 (header_row_fill, required_header_fill, required_font_color)。
        提供安全的默认值以避免配置缺失导致错误。
        """
        cfg_candidates = (
            self.settings.get('ui'),
            self.settings.get('visual'),
            self.settings.get('setting_config'),
            self.settings.get('colors'),
            {}
        )
        cfg = {}
        for c in cfg_candidates:
            if isinstance(c, dict) and c:
                cfg = c
                break

        header = cfg.get('header_row_fill', cfg.get('header_fill', '#D9EEFF'))
        required = cfg.get('required_header_fill', cfg.get('required_fill', '#0B66C3'))
        req_font = cfg.get('required_font_color', '#FFFFFF')
        # 规范化带或不带#的颜色
        def norm(c): return (c or '').lstrip('#')[:6].upper() if c else ''
        return ('#' + norm(header), '#' + norm(required), '#' + norm(req_font))

    def apply_excel_formatting(self, workbook, worksheet, df,auto_enter=False):
        """对Excel应用列宽、表头格式等美化"""
        try:
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter
        except Exception as e:
            print(f"无法导入openpyxl依赖:{e}\n 注意如果要应用excel格式，你需要安装openpyxl依赖包")
            return
        
        # 获取样式（优先使用 df.attrs 中保存的初始化样式）
        hs = df.attrs.get('header_styles', None) if hasattr(df, 'attrs') else None
        if hs:
            header_row_color = hs.get('header_row_fill', '#D9EEFF')
            required_color = hs.get('required_header_fill', '#0B66C3')
            required_font_color = hs.get('required_font_color', '#FFFFFF')
        else:
            header_row_color, required_color, required_font_color = self.get_header_styles()

        header_fill = PatternFill(start_color=header_row_color.lstrip('#'), 
                                end_color=header_row_color.lstrip('#'), 
                                fill_type="solid")
        header_font = Font(bold=True)
        required_fill = PatternFill(start_color=required_color.lstrip('#'), 
                                end_color=required_color.lstrip('#'), 
                                fill_type="solid")
        required_font = Font(color=required_font_color.lstrip('#'), bold=True)

        # 第一行表头样式
        for i, col in enumerate(df.columns):
            col_letter = get_column_letter(i + 1)
            cell = worksheet[f"{col_letter}1"]
            cell.fill = header_fill
            cell.font = header_font

        # 对必填列的表头应用深色标注
        required_vars = [t['table_name'] for t in self.config.get_required_tags()]
        for i, col in enumerate(df.columns):
            if col in required_vars:
                col_letter = get_column_letter(i + 1)
                cell = worksheet[f"{col_letter}1"]
                cell.fill = required_fill
                cell.font = required_font
        # 强制所有数据单元格为文本格式 
        # min_row=2 表示跳过表头，从数据行开始
        if worksheet.max_row >= 2:
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                         min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    # '@' 是 Excel 的文本格式代码
                    cell.number_format = '@'

        # 设置列宽 - 修复列宽计算问题
        for column in df.columns:
            column_letter = get_column_letter(df.columns.get_loc(column) + 1)
            
            # 计算合适的列宽
            if df.empty:
                # 如果DataFrame为空，使用列名的长度
                column_length = len(str(column))
            else:
                # 如果有数据，计算列名和列中数据的最大长度
                try:
                    import pandas as pd
                    # 确保数据是字符串类型
                    column_data = df[column].astype(str)
                    max_data_length = column_data.map(len).max()
                    column_length = max(len(str(column)), max_data_length)
                except:
                    column_length = len(str(column))
            
            # 确保最小列宽为15，最大不超过50
            adjusted_width = max(min(column_length + 3, 50), 15)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 确保所有行都可见（解除隐藏）
        worksheet.sheet_format.defaultRowHeight = 15
        
        # 设置单元格自动换行格式
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                    min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = cell.alignment.copy(wrapText=auto_enter)

# 创建全局单例
_update_file_utils_instance = None

def get_update_file_utils():
    """获取更新文件工具单例"""
    global _update_file_utils_instance
    if _update_file_utils_instance is None:
        _update_file_utils_instance = UpdateFileUtils()
    return _update_file_utils_instance