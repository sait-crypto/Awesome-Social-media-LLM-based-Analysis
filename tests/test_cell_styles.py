"""Debug Excel file after save to see cell fills"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from src.core.database_manager import DatabaseManager
from openpyxl import load_workbook
import pandas as pd

def test_cell_styles():
    """Test cell styles in saved Excel"""
    
    db = DatabaseManager()
    excel_path = db.core_excel_path
    
    print(f"=== 检查 Excel 文件单元格样式 ===")
    print(f"文件: {excel_path}")
    
    # Load workbook directly
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Get invalid_fields column
    df = pd.read_excel(excel_path, nrows=1)  # 只读表头
    invalid_col_name = "invalid fields"
    
    if invalid_col_name in df.columns:
        col_idx = df.columns.get_loc(invalid_col_name)
        col_letter = chr(65 + col_idx)
        
        print(f"\ninvalid_fields 列: {invalid_col_name}")
        print(f"列索引: {col_idx}, 列字母: {col_letter}")
        
        print(f"\n检查前30行单元格:")
        for row_idx in range(2, min(32, ws.max_row + 1)):
            cell = ws[f"{col_letter}{row_idx}"]
            
            fill_info = "无填充"
            if cell.fill:
                start_color = cell.fill.start_color
                if start_color:
                    if hasattr(start_color, 'rgb'):
                        fill_info = f"RGB={start_color.rgb}"
                    elif hasattr(start_color, 'index'):
                        fill_info = f"Index={start_color.index}"
                    else:
                        fill_info = f"填充={start_color}"
                else:
                    fill_info = "填充对象存在但无颜色"
            
            print(f"  Row {row_idx}: Value={cell.value}, Fill={fill_info}, Number Format={cell.number_format}")
            
            # 检查是否是需要上色的单元格
            if cell.value and cell.value not in ['0', '', None, 'nan', 'false']:
                print(f"    ^ 这个单元格应该被标红!")

if __name__ == "__main__":
    test_cell_styles()
